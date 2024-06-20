import tkinter as tk
from tkinter import filedialog
import shutil
import os
import csv
from openpyxl import Workbook, load_workbook

target_folder = r"C:\Users\Z3R0\Desktop\trial"
destination_file = r"C:\Users\Z3R0\Desktop\trial\consolidated.xlsx"

def browse_files():
    filenames = filedialog.askopenfilenames(initialdir="/",
                                            title="Select Files",
                                            filetypes=(("Text files", "*.txt*"),
                                                       ("CSV files", "*.csv"),
                                                       ("Excel files", "*.xlsx"),
                                                       ("all files", "*.*")))
    if filenames:
        for filename in filenames:
            upload_file(filename)

def upload_file(filepath):
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)
    try:
        if filepath.lower().endswith('.csv'):
            xlsx_filepath = os.path.join(target_folder, os.path.basename(filepath)[:-4] + '.xlsx')
            convert_csv_to_xlsx(filepath, xlsx_filepath)
            filepath = xlsx_filepath
        shutil.copy(filepath, target_folder)
    except Exception as e:
        print(f"Error uploading {filepath}: {e}")

def convert_csv_to_xlsx(csv_filepath, xlsx_filepath):
    workbook = Workbook()
    sheet = workbook.active
    with open(csv_filepath, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            sheet.append(row)
    workbook.save(xlsx_filepath)

def open_window():
    global new_window
    new_window = tk.Toplevel(root)
    new_window.title("New Window")

    global col_no_label, col_no_entry, col_name_entries
    col_no_label = tk.Label(new_window, text="Enter the no. of columns:", fg="black")
    col_no_label.grid(column=1, row=1, padx=10, pady=10)

    col_no_entry = tk.Entry(new_window)
    col_no_entry.grid(column=2, row=1, padx=10, pady=10)

    col_name_entries = []

    col_name_label = tk.Label(new_window, text="Enter column name(s):", fg="black")
    col_name_label.grid(column=1, row=2, padx=10, pady=10)

    button_ok = tk.Button(new_window,
                          text="OK",
                          borderwidth=5,
                          command=entrybox)
    button_ok.grid(column=3, row=1, padx=10, pady=10)

def entrybox():
    global new_window, col_name_entries
    for entry in col_name_entries:
        entry.destroy()
    col_name_entries.clear()

    try:
        num_columns = int(col_no_entry.get().strip())
        for i in range(num_columns):
            col_name_entry = tk.Entry(new_window)
            col_name_entry.grid(column=1, row=i+3, padx=10, pady=10)
            col_name_entries.append(col_name_entry)
        submit = tk.Button(new_window,
                           text="Submit",
                           command=lambda: collect_col_names(col_name_entries))
        submit.grid(column=1, row=num_columns+3, padx=10, pady=10)
    except ValueError:
        print("Please enter a valid integer for the number of columns.")

def collect_col_names(entries):
    global col_names_list
    col_names_list = [entry.get().strip() for entry in entries]
    print("Column Names:", col_names_list)
    check_columns_in_files()

def check_columns_in_files():
    if not os.path.exists(destination_file):
        dest_wb = Workbook()
        dest_wb.save(destination_file)

    dest_wb = load_workbook(destination_file)
    dest_sheet = dest_wb.active

    # Dictionary to store column indices in the destination file
    column_indices = {}

    # Dictionary to store column values
    column_values = {col_name: [] for col_name in col_names_list}

    # Iterate over files in the target folder
    files = os.listdir(target_folder)
    for file in files:
        if file.endswith('.xlsx') and not file.startswith('~$'):
            xlsx_filepath = os.path.join(target_folder, file)
            try:
                workbook = load_workbook(xlsx_filepath)
                sheet = workbook.active
                headers = [cell.value.strip() for cell in sheet[1]]  # Assuming first row is the header

                # Iterate over column names
                for col_name in col_names_list:
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        if col_name not in column_indices:
                            # Add column header to destination file if not present
                            column_indices[col_name] = len(column_indices) + 1
                            dest_sheet.cell(row=1, column=column_indices[col_name], value=col_name)

                        # Iterate over rows and append data to column_values dictionary
                        for row_idx, cell in enumerate(sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2), start=2):
                            column_values[col_name].append(cell[0].value)

                print(f"Processed columns from {file}.")
            except Exception as e:
                print(f"Error reading {file}: {e}")

    if column_values and any(column_values.values()):
        max_rows = max(len(column_values[col_name]) for col_name in column_values)
        for col_name, values in column_values.items():
            for row_idx, value in enumerate(values, start=2):
                dest_sheet.cell(row=row_idx, column=column_indices[col_name], value=value)

        dest_wb.save(destination_file)
        print(f"Data saved to {destination_file}.")
    else:
        print("No matching columns found in the provided files.")

def view_files():
    if os.path.exists(target_folder):
        files = os.listdir(target_folder)
        if files:
            file_list = "\n".join(files)
            global new_windoww
            new_windoww = tk.Toplevel(root)
            new_windoww.title("New Window")
            labell = tk.Label(new_windoww, text=file_list, fg="black")
            labell.pack()
        else:
            label_file_explorer.config(text="No files in the folder.")
    else:
        label_file_explorer.config(text="Target folder does not exist.")

root = tk.Tk()
root.title('File Explorer')
root.geometry("400x300")

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

frame = tk.Frame(root)
frame.grid(column=0, row=0, padx=20, pady=20)
frame.columnconfigure(0, weight=1)
frame.rowconfigure(0, weight=1)

label_file_explorer = tk.Label(frame, text="File Explorer using Tkinter",
                               width=50, height=2, fg="black")
label_file_explorer.grid(column=0, row=0, padx=10, pady=10)

button_explore = tk.Button(frame,
                           text="Browse Files",
                           command=browse_files)
button_explore.grid(column=0, row=1, padx=10, pady=10)

button_view = tk.Button(frame,
                        text="View Files",
                        command=view_files)
button_view.grid(column=0, row=2, padx=10, pady=10)

button_exit = tk.Button(frame,
                        text="Exit",
                        command=root.quit)
button_exit.grid(column=0, row=3, padx=10, pady=10)

button_read = tk.Button(frame,
                        text="Read",
                        command=open_window)
button_read.grid(column=0, row=4, padx=10, pady=10)

root.mainloop()